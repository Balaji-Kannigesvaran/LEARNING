# DIABLO CONNECTION FOR SHS WITH FULL CAPACITY
import math as ma
import numpy as np
import pandas as pd

from openpyxl import Workbook
workbook = Workbook()
sheet = workbook.active

# define function
def ceiling(x,s):
    if (x%s==0):return x
    else:return (s+1)*(x/s)
def floor(x,s):
    if (x%s==0):return x
    else:return (s)*(x/s)

# Section dimensions
D=180                                                                          # Total depth of section
B=180                                                                          # Width of section
tf=4                                                                           # Thickness of flange
tw=tf                                                                          # Thickness of web
d=D-2*tf                                                                       # Clear depth
b=B-2*tw                                                                       # Clear width
L=2900                                                                         # Member Length

#Reading Excel
xls=pd.ExcelFile('D:\Balaji work\PROGRAM\DIABLO\Box Section Diablo.xlsx')
df=pd.read_excel(xls,"Tables")

# Section Properties
A=(B*D)-(b*d)                                                                  # Area of Cross Section
W=A*7850/1000000                                                               # Wt/m of the section
Avd=(A*D)/(B+D)                                                                # Shear area along Depth            
Avb=(A*B)/(B+D)                                                                # Shear area along Width
Ixx=B*D**3/12-b*d**3/12                                                        # Moment of Inertia
Iyy=D*B**3/12-d*b**3/12                                                        # Moment of Inertia
Rxx=ma.sqrt(Ixx/A)                                                             # Radius of gyration
Ryy=ma.sqrt(Iyy/A)                                                             # Radius of gyration
BC='a'                                                                         # Buckling Class
df1=df.iloc[21:23,4:8]                                                         # Reading Buckling class table
# Imperfection factor by converting dataframe(df) to list
#df1_list=df1.values.tolist()
#for i in range (0,4):                                
#   if BC==df1_list[0][i]:
#        IF=df1_list[1][i]
# Imperfection factor by converting dataframe(df) to array
df1_array=df1.to_numpy()
IF=df1_array[1][np.where(df1_array[0]==BC)]

#Strength properties
E=200000
SG='E350'                                                                               # Steel Grade
df2=df.iloc[5:18,2:7]                                                                   # Reading steel grade table
df2_array=df2.to_numpy()
fy_sec=df2_array[np.where(df2_array[:,0]==SG)][0][2 if tf<20 else 3 if tf<=40 else 4]   # Yield strength of section
fu_sec=df2_array[np.where(df2_array[:,0]==SG)][0][1]                                    # Ultimate strength of section

#Partial Safety Factors for materials
gamma_mo=1.1                                                                   # Resistance governed by yielding
gamma_ml=1.25                                                                  # Resistance governed by ultimate stress
gamma_mf=1.25                                                                  # Resistance governed by bolt friction type
gamma_mb=1.25                                                                  # Resistance governed by bolt bearing type
gamma_mw=1.25                                                                  # Resistance governed by shop fabrication

#Tension Capacity of Section
Td=A*fy_sec/gamma_mo/1000
#Shear Capacity of Section  
Vd=(max(Avd,Avb)*fy_sec)/(ma.sqrt(3)*gamma_mo)/1000
#Compression Capacity of the section
k=1                                                                            # Slenderness coefficient
SR=k*L/min(Rxx,Ryy)                                                            # Slenderness Ratio
fcc=ma.pi**2*E/(SR**2)                                                         # Euler Buckling Stress
lamda=ma.sqrt(fy_sec/fcc)                                                      # Non-dimensional efective slenderness ratio
phi=0.5*(1+IF*(lamda-0.2)+(lamda**2))
Xeta=1/(phi+(phi**2-lamda**2)**0.5)                                            # Stress Reduction Factor
fcd=min(Xeta*fy_sec/gamma_mo,fy_sec/gamma_mo)                                  # Design Compressive stress
Pd=fcd*A/1000                                                                  # Compression Capacity
#print(Td,Vd,Pd,sep='\n')

#Bolt Details
Gr=8.8                                                                         # Grade of bolt
Nc=4                                                                           # No of corner bolts
Nh=2                                                                           # No of Horizontal bolts
Nv=2                                                                           # No of Vertical bolts
Nprov=Nc+Nh+Nv                                                                 # No. of bolts provided
tep_array=[3,4,6,8,10,12,16,20,25,32,36,40,50,60]                              # End plate thickness
EPSG='E350'                                                                    # End Plate Steel Grade
fyp=df2_array[np.where(df2_array[:,0]==SG)][0][2 if tf<20 else 3 if tf<=40 else 4]   # Yield strength of End plate
fup=df2_array[np.where(df2_array[:,0]==SG)][0][1]                                    # Ultimate strength of End Plate                       
db_array=np.array([5,6,8,10,12,16,20,24,30,36,42,48,56,64])                    # Dia of bolt array
re = np.empty((0,8))
i = 0
for db in range (0,len(db_array)):
    db=db_array[db]                                                            # Dia of bolt
    dh=db+1 if db<16 else db+2 if db<=24 else db+3                             # Dia of hole
    df3=df.iloc[68:79,2:7]                                                     # Reading steel grade table
    df3_array=df3.to_numpy()
    fyb=df3_array[np.amin(np.where(df3_array[:,0]==Gr))+(1 if db>16 else 0)][2]# Yield strength of bolt
    fub=df3_array[np.amin(np.where(df3_array[:,0]==Gr))+(1 if db>16 else 0)][3]# Ultimate strength of bolt
    Pmin=2.5*db                                                                # Min. Pitch Required
    emin=1.5*dh                                                                # Min. Edge Dist Required
    nn=1
    ns=0
    Anb=(ma.pi/4)*db**2
    Asb=0.78*Anb    
    Pphi=ceiling(Pmin,5)                                                       # Horizontal Pitch Provided
    Ppvi=ceiling(Pmin,5)                                                       # Vertical Pitch Provided
    for Pph in range (int(Pphi),B,5):
        eph=(B-((((Nh+Nc)/2)-1)*Pph))/2                                        # Horizontal Edge Provided
        if eph<emin or Pph<Pmin:
            continue
        else:
            # Bolt Design
            # Tension capacity of bolt
            Tnbh=min(0.9*fub*Anb,fyb*Asb*gamma_mb/gamma_mo)
            Tdbh=Tnbh/gamma_mb/1000
            # Shear capacity of bolt
            Vdsbh=(fub/(ma.sqrt(3)*gamma_mb*1000))*(nn*Anb+ns*Asb)
            # Bearing capacity of Plate
            for tep in range (0,len(tep_array)):
                tep=tep_array[tep]
                kbh=min(eph/(3*dh),(Pph/(3*dh))-0.25,fub/fup,1)
                Vdpbh=2.5*kbh*db*tep*fup/1000
                # print("HORIZONTAL",db,Pmin,Pph,emin,eph,((((Nh+Nc)/2)-1)*Pph)+(2*eph),'\n')
                # print(Tdbh,Vdsbh,kbh,Vdpbh,'\n')
                re = np.append(re,np.round(np.array([db,Pmin,Pph,emin,eph,Tdbh,Vdsbh,Vdpbh]),1))
                sheet.cell(i+1,1,db)
                sheet.cell(i+1,2,Pmin)
                sheet.cell(i+1,3,Pph)
                sheet.cell(i+1,4,emin)
                sheet.cell(i+1,5,eph)
                sheet.cell(i+1,6,Tdbh)
                sheet.cell(i+1,7,Vdsbh)
                sheet.cell(i+1,8,Vdpbh)
                i = i+1
                Vdsh=min(Vdsbh,Vdpbh)
                Nbhr=Vd/Vdsh
    for Ppv in range (int(Ppvi),D,5):
        epv=(D-((((Nv+Nc)/2)-1)*Ppv))/2                                        # Vertical Edge Provided
        if epv<emin or Ppv<Pmin:
            continue
        else:
            # Bolt Design
            # Tension capacity of bolt
            Tnbv=min(0.9*fub*Anb,fyb*Asb*gamma_mb/gamma_mo)
            Tdbv=Tnbv/gamma_mb/1000
            # Shear capacity of bolt
            Vdsbv=(fub/(ma.sqrt(3)*gamma_mb*1000))*(nn*Anb+ns*Asb)
            # Bearing capacity of Plate
            for tep in range (0,len(tep_array)):
                tep=tep_array[tep]
                kbv=min(epv/(3*dh),(Ppv/(3*dh))-0.25,fub/fup,1)
                Vdpbv=2.5*kbv*db*tep*fup/1000
                # print("VERTICAL",db,Pmin,Ppv,emin,epv,((((Nv+Nc)/2)-1)*Ppv)+(2*epv),'\n')
                # print(Tdbv,Vdsbv,kbv,Vdpbv,'\n')
                Vdsv=min(Vdsbv,Vdpbv)
                Nbhr=Vd/Vdsv
#print(re)
re1 = np.reshape(re,(int(np.size(re)/8),8))
#print(np.reshape(re,(int(np.size(re)/8),8)))

re2=[]
re2=[val for re2 in re1 for val in re2]
df = pd. DataFrame(re1, columns=['db', 'Pmin','P', 'emin','e', 'Tdb','Vdsb', 'Vdpb'])
print(df)
df.to_excel('D:\\Balaji work\\PROGRAM\\DIABLO\\test.xlsx',index=False)
# workbook.save(filename="array.xlsx")

# print(re1[:,2:])


#print(dh,Pmin,emin,fyb,fub,Anb,Asb,sep='\n')

#End Plate Details
Dep=D                                              #Depth of End Plate
Bep=B                                              #Width of End Plate
Tep=20                                             #Thickness of End Plate
Bop=30                                             #Width of the opening   
Aop=ma.pi/4*Bop**2                                 #Area of opening
q=Pd*1000/((Dep*Bep)-Aop)                          #Design Pressure


#print(Aop,q,sep='\n')

